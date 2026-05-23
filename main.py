from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import time

data = []

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://news.ycombinator.com/")
time.sleep(2)

post_title = driver.find_elements(By.CLASS_NAME, "titleline")
score = driver.find_elements(By.CLASS_NAME, "score")
links = driver.find_elements(By.CLASS_NAME, "titleline")
user_names = driver.find_elements(By.CLASS_NAME, "hnuser")
times = driver.find_elements(By.CLASS_NAME, "age")

for p, s, l, u, t in zip(post_title, score, links, user_names, times):
    a_tag = l.find_element(By.TAG_NAME, "a")
    link_actual = a_tag.get_attribute("href")
    data.append({
        "Title": p.text,
        "Score": s.text,
        "Link": link_actual,
        "Username": u.text,
        "Time": t.text
    })

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ── LOAD DATA ─────────────────────────────────────────────────
df = pd.read_excel("/Users/syedhasan/Desktop/PycharmProjects/Amazon scpraper/hackernews.xlsx")

# ── SETUP WORKBOOK ────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Hacker News — Top Posts"

# ── COLORS ────────────────────────────────────────────────────
ORANGE       = "FF6600"   # HN brand orange
DARK_BG      = "1A1A2E"   # dark header background
WHITE        = "FFFFFF"
LIGHT_ROW    = "F8F9FA"   # light grey alternating row
DARK_ROW     = "EAEEF2"   # slightly darker alternating row
BORDER_COLOR = "D0D7DE"
TEXT_DARK    = "24292F"
TEXT_LINK    = "0969DA"   # blue for links

# ── BORDER ────────────────────────────────────────────────────
thin = Side(style="thin", color=BORDER_COLOR)
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── TITLE ROW ─────────────────────────────────────────────────
ws.merge_cells("A1:E1")
title_cell = ws["A1"]
title_cell.value = "🔶  Hacker News — Daily Tech Intelligence Report"
title_cell.font = Font(name="Arial", bold=True, size=14, color=WHITE)
title_cell.fill = PatternFill("solid", start_color=DARK_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── SUBTITLE ──────────────────────────────────────────────────
ws.merge_cells("A2:E2")
sub_cell = ws["A2"]
sub_cell.value = f"Top {len(df)} posts scraped from news.ycombinator.com"
sub_cell.font = Font(name="Arial", italic=True, size=10, color="6E7781")
sub_cell.fill = PatternFill("solid", start_color="F0F0F5")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# ── HEADER ROW ────────────────────────────────────────────────
headers = ["Title", "Score", "Link", "Username", "Time Posted"]
header_row = 3

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col_idx)
    cell.value = header
    cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    cell.fill = PatternFill("solid", start_color=ORANGE)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = cell_border

ws.row_dimensions[header_row].height = 28

# ── DATA ROWS ─────────────────────────────────────────────────
for row_idx, row in df.iterrows():
    excel_row = row_idx + 4
    is_even = row_idx % 2 == 0
    bg_color = LIGHT_ROW if is_even else DARK_ROW

    values = [
        row.get("Title", "N/A"),
        row.get("Score", "N/A"),
        row.get("Link", "N/A"),
        row.get("Username", "N/A"),
        row.get("Time", "N/A"),
    ]

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.value = value
        cell.fill = PatternFill("solid", start_color=bg_color)
        cell.border = cell_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

        # title — left aligned, dark text
        if col_idx == 1:
            cell.font = Font(name="Arial", size=10, color=TEXT_DARK, bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # score — centered, orange bold
        elif col_idx == 2:
            cell.font = Font(name="Arial", size=10, color=ORANGE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # link — blue, left aligned
        elif col_idx == 3:
            cell.font = Font(name="Arial", size=9, color=TEXT_LINK)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # username — centered
        elif col_idx == 4:
            cell.font = Font(name="Arial", size=10, color=TEXT_DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # time — centered, grey italic
        elif col_idx == 5:
            cell.font = Font(name="Arial", size=10, color="6E7781", italic=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # auto row height based on title length
    title_len = len(str(values[0]))
    if title_len > 80:
        ws.row_dimensions[excel_row].height = 42
    elif title_len > 50:
        ws.row_dimensions[excel_row].height = 32
    else:
        ws.row_dimensions[excel_row].height = 22

# ── COLUMN WIDTHS ─────────────────────────────────────────────
col_widths = [52, 10, 55, 16, 16]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── FREEZE HEADER ─────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── SAVE ──────────────────────────────────────────────────────
output_path = "/Users/syedhasan/Desktop/PycharmProjects/Amazon scpraper/hackernews_report.xlsx"
wb.save(output_path)
print(f"Done. Report saved to {output_path}")

driver.quit()